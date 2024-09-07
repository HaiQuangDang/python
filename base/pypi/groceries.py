import openpyxl as xl
workbook = xl.load_workbook("groceries.xlsx")
sheet = workbook["Sheet1"]
cell1 = sheet.cell(1, 2)

for row in range(2, sheet.max_row + 1):
    cell = sheet.cell(row, 3)
    if not cell.value == None:
        reduce_price = cell.value * 0.9
        reduce_price_cell = sheet.cell(row, 4)
        reduce_price_cell.value = reduce_price
    else:
        break

workbook.save("groceries2.xlsx")



